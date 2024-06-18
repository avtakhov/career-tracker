import io

import sqladmin
import wtforms
from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.db.base import async_session
from app.services.admin.views.group import insert_users, insert_users_groups
from app.services.bot.models.group import Group


class GroupForm(wtforms.Form):
    group_name = wtforms.StringField("Имя группы", render_kw={"class": "form-control"})
    xlsx = wtforms.FileField("Excel файл", render_kw={"class": "form-control"})


class GroupAdmin(sqladmin.ModelView, model=Group):
    icon = "fa-solid fa-users"
    name_plural = "Группы"
    column_list = [Group.group_name]
    form = GroupForm
    form_include_pk = True
    can_export = False

    async def after_model_change(self, data, group: Group, *args):
        xlsx: UploadFile = data.get("xlsx")
        if xlsx.size == 0:
            return

        sheet: Worksheet = load_workbook(
            io.BytesIO(xlsx.file.read()),
            read_only=True
        ).active

        xlsx.file.close()

        users = [
            {'full_name': row[0].strip(), 'telegram_username': row[1].strip()}
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[0] and row[1]
        ]
        async with async_session() as session:
            await insert_users(users, session)
            await insert_users_groups(users, group.group_name, session)
            await session.commit()
